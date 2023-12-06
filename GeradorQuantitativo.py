import arcpy
import os.path
import pandas as pd
import numpy as np
import openpyxl as op

arcpy.env.overwriteOutput = True
arcpy.env.addOutputsToMap = False


lt_inteira = arcpy.GetParameterAsText(0)#linha de transmissão previa
circuito_duplo = arcpy.GetParameterAsText(2)#insira a coluna que divide os circuitos
vert_inicial = arcpy.GetParameterAsText(4)#a linha começa no vertice 01 ou 02?
fx_interesse = arcpy.GetParameterAsText(5)#insira a faixa de interesse
bdgis = arcpy.GetParameterAsText(6)#insira o endereço do banco de dados GIS
temas_extra = arcpy.GetParameterAsText(11)#multiple values (separados por vírgula) de temas que não estão no banco de dados GIS
fields_extras = arcpy.GetParameterAsText(13)#multiple values (separados por vírgula) de campos extras que não estão no banco de dados GIS
gdb_path = arcpy.GetParameterAsText(7)#caminho do geodatabase final, que terá os temas que estão no banco de dados GIS
pasta_quantitativo = arcpy.GetParameterAsText(9)#pasta onde serão salvos os arquivos excel
atualizar = arcpy.GetParameterAsText(8)
atualizar_vao = arcpy.GetParameterAsText(3)
junkspace = arcpy.GetParameterAsText(10)
geodesic = arcpy.GetParameterAsText(1)
fields_tema_extra = arcpy.GetParameterAsText(12)
divisao_estadual = r'C:\Users\anderson.souza\Downloads\BR_UF_2022\BR_UF_2022.shp'
arcpy.env.workspace = gdb_path
feature_datasets = arcpy.ListDatasets()

if 'Quantitativo' not in feature_datasets:
    arcpy.CreateFeatureDataset_management(gdb_path, 'Quantitativo', arcpy.Describe(lt_inteira).spatialReference)
else:
    pass
gdb_quantitativo = os.path.join(gdb_path, 'Quantitativo')
#verifica se a pasta junkspace esta vazia, se não estiver, deleta tudo
if os.listdir(junkspace) != []:
    for file in os.listdir(junkspace):
        os.remove(os.path.join(junkspace, file))



def criavao(shape_lt, fx_interesse,vert_inicial):
    # Split at Vertices
    output_split = r'in_memory\"SplitVertices"'
    arcpy.SplitLine_management(shape_lt, output_split)

    # Criar novo campo "Sequencial"
    campo_sequencial = 'Sequencial'
    arcpy.management.AddField(output_split, campo_sequencial, 'LONG')
    if vert_inicial == 'Inicia no Vértice 0':
        vert_inicial = 0
    elif vert_inicial == 'Inicia no Vértice 1':
        vert_inicial = 1
    # Enumerar os registros sequencialmente
    with arcpy.da.UpdateCursor(output_split, campo_sequencial) as cursor:
        for i, row in enumerate(cursor,start=vert_inicial):
            row[0] = i
            cursor.updateRow(row)

    # Criar novo campo "Vertices" e preenchê-lo com a enumeração
    campo_vertices = 'Vertices'
    arcpy.management.AddField(output_split, campo_vertices, 'TEXT', field_length=20)

    with arcpy.da.UpdateCursor(output_split, [campo_sequencial, campo_vertices]) as update_cursor:
        for row in update_cursor:
            sequencia = row[0]
            if sequencia > 0:
                valor_anterior = f"V{sequencia:02d}"
                valor_atual = f"V{sequencia + 1:02d}"
                row[1] = f"{valor_anterior}-{valor_atual}"
            else:
                row[1] = f"V{sequencia:02d}-V{sequencia + 1:02d}"
            update_cursor.updateRow(row)

    # Salvar a feature class no geodatabase de saída com o nome "Diretriz_Gerada"
    try:
        output_diretriz_gerada = arcpy.CopyFeatures_management(output_split, os.path.join(gdb_path,'Dados_Caruso','Vao_LT'))
    except:
        #delete feature class
        arcpy.management.Delete(os.path.join(gdb_path,'Dados_Caruso','Vao_LT'))
        output_diretriz_gerada = arcpy.CopyFeatures_management(output_split, os.path.join(gdb_path,'Dados_Caruso','Vao_LT'))

    #largura da faixa de servidao em metros
    fs_distancia = str(float(fx_interesse)/2)+' Meters'

    # Caminho completo para o arquivo shapefile do buffer
    output_buffer_saida = r'in_memory\"faixa_servidao_buffer"'

    # Salvar o buffer como shapefile
    arcpy.Buffer_analysis(output_diretriz_gerada, output_buffer_saida, fs_distancia)
    spatial_reference = arcpy.Describe(shape_lt).spatialReference
    # Criar a feature class em memória
    no_over = arcpy.management.CreateFeatureclass(
        "in_memory",  # Caminho "in_memory" indica que a feature class será criada em memória
        "Faixa_Servidao",  # Nome da feature class
        "POLYGON",  # Tipo de geometria (neste caso, polígono)
        None,  # Template - pode ser None para não usar um template
        "DISABLED",  # Habilitar M (Medidas) - neste caso, está desabilitado
        "DISABLED",  # Habilitar Z (Elevação) - neste caso, está desabilitado
        spatial_reference,
        '',  # Configurações do cluster - vazio neste caso
        0,  # Recursos do tamanho do arquivo - 0 neste caso para uso padrão
        0,  # Armazenamento do limite do arquivo - 0 neste caso para uso padrão
        0,  # Armazenamento do limite de espaço único - 0 neste caso para uso padrão
    )

    # Criar os campos "Sequencial" e "Vertices" na feature class "no_over"
    campo_sequencial = 'Sequencial'
    campo_vertices = 'Vertices'
    arcpy.management.AddField(no_over, campo_sequencial, 'LONG')
    arcpy.management.AddField(no_over, campo_vertices, 'TEXT', field_length=20)

    # Copiar os valores dos campos "Sequencial" e "Vertices" da feature class "Diretriz_Gerada" para a feature class "no_over"
    campos_copiar = [campo_sequencial, campo_vertices]
    with arcpy.da.UpdateCursor(no_over, campos_copiar) as update_cursor:
        for row in update_cursor:
            seq_valor = row[0]  # Valor do campo Sequencial na feature class "Diretriz_Gerada"
            vert_valor = row[1]  # Valor do campo Vertices na feature class "Diretriz_Gerada"
            # Copiar os valores para a feature class "no_over"
            row[0] = seq_valor
            row[1] = vert_valor
            update_cursor.updateRow(row)


    # Adicionar novo campo "Circuito" na feature class "no_over"
    if circuito_duplo == 'true':
        campo_circuito = 'Circuito'
        arcpy.management.AddField(no_over, campo_circuito, 'TEXT', field_length=50)

        # Preencher o campo "Circuito" na faixa de servidão com os valores correspondentes da LT
        with arcpy.da.SearchCursor(output_diretriz_gerada, ['Sequencial', 'Circuito']) as search_cursor:
            for row_search in search_cursor:
                sequencial = row_search[0]
                circuito_lt = row_search[1]

                # Atualizar o campo "Circuito" na faixa de servidão com o valor correspondente da LT
                with arcpy.da.UpdateCursor(no_over, ['Sequencial', campo_circuito]) as update_cursor:
                    for row_update in update_cursor:
                        if row_update[0] == sequencial:
                            row_update[1] = circuito_lt
                            update_cursor.updateRow(row_update)
                            break

    sl_fx_serv = arcpy.management.SelectLayerByAttribute(in_layer_or_view=output_buffer_saida, selection_type='NEW_SELECTION', where_clause='Sequencial=0')
    feature_copiada = os.path.join(gdb_path,'Dados_Caruso','feature_copiada')
    copyfzero = arcpy.management.CopyFeatures(sl_fx_serv, feature_copiada, '', None, None, None)
    no_over_final = arcpy.management.Append(inputs=copyfzero, target=no_over, schema_type='NO_TEST')
    #deleta a feature class copiada
    arcpy.management.Delete(feature_copiada)

    cursor = arcpy.SearchCursor(output_buffer_saida)

    for row in cursor:
        arcpy.env.workspace = junkspace
        vao_row = row.getValue('Sequencial')
        sl_getcount = arcpy.management.SelectLayerByAttribute(in_layer_or_view=output_buffer_saida, selection_type='NEW_SELECTION', where_clause=f'"Sequencial"={vao_row}')
        copy_getcount = arcpy.management.CopyFeatures(sl_getcount, "teste2.shp", '', None, None, None) 
        erase_getcount = arcpy.analysis.Erase(in_features=copy_getcount, erase_features=no_over_final, out_feature_class=f'in_memory\\erase_copy_{vao_row}')

        # Create a FieldMappings object and add the necessary field mappings
        field_mappings = arcpy.FieldMappings()
        field_mappings.addTable(no_over_final)
        field_mappings.addTable(erase_getcount)

        # Append the features from erase_getcount to no_over_final using the defined field mappings
        arcpy.management.Append(inputs=erase_getcount, target=no_over_final, schema_type='NO_TEST', field_mapping=field_mappings)
    # Salvar a feature class no geodatabase de saída com nome "Faixa_Servidao"
    try: 
        output_faixa_servidao = arcpy.CopyFeatures_management(no_over, os.path.join(gdb_path,'Dados_Caruso','Vao_FxInteresse'))
    except:
        #delete feature class
        arcpy.management.Delete(os.path.join(gdb_path,'Dados_Caruso','Vao_FxInteresse'))
        output_faixa_servidao = arcpy.CopyFeatures_management(no_over, os.path.join(gdb_path,'Dados_Caruso','Vao_FxInteresse'))
    arcpy.AddMessage('LT e Faixa de Interesse segmentadas geradas com sucesso')
    return output_diretriz_gerada, output_faixa_servidao

def project(gdb,temas_extra):
    arcpy.env.workspace = gdb
    feature_datasets = arcpy.ListDatasets()
    if 'Temas' not in feature_datasets:
        arcpy.CreateFeatureDataset_management(gdb, 'Temas', arcpy.Describe(lt_inteira).spatialReference)
    arcpy.env.workspace = bdgis
    feature_classes = arcpy.ListFeatureClasses()
    if temas_extra != '':
        #repair geometry
        temas_extra_name = os.path.basename(temas_extra).replace('.shp','')
        if 'UF' in fields_extras:
            fc_intersect = arcpy.analysis.Identity(temas_extra, divisao_estadual, os.path.join(junkspace,fr'div_{temas_extra_name}'))
            fc_layer = arcpy.management.MakeFeatureLayer(fc_intersect, 'fc_layer')
            fc_select = arcpy.management.SelectLayerByLocation(in_layer=fc_layer, overlap_type='WITHIN_A_DISTANCE', select_features=lt_inteira, search_distance='60000 Meters')
            arcpy.conversion.FeatureClassToFeatureClass(fc_select, os.path.join(gdb, 'Temas'),temas_extra_name)
        elif 'UF' not in fields_extras:
            fc_layer = arcpy.management.MakeFeatureLayer(temas_extra, 'fc_layer')
            fc_select = arcpy.management.SelectLayerByLocation(in_layer=fc_layer, overlap_type='WITHIN_A_DISTANCE', select_features=lt_inteira, search_distance='60000 Meters')
            arcpy.conversion.FeatureClassToFeatureClass(fc_select, os.path.join(gdb, 'Temas'),temas_extra_name)
    else:
        for fc in feature_classes:
            arcpy.management.RepairGeometry(fc)
            filename = os.path.basename(fc)
            #intersect entre a fc e a divisão estadual
            if 'UF' in fields(fc):
                fc_intersect = arcpy.analysis.Identity(fc, divisao_estadual, os.path.join(junkspace,fr'div_{filename}'))
                #select by location para selecionar apenas os poligonos que estao em um raio de 50km da lt
                fc_layer = arcpy.management.MakeFeatureLayer(fc_intersect, 'fc_layer')
                fc_select = arcpy.management.SelectLayerByLocation(in_layer=fc_layer, overlap_type='WITHIN_A_DISTANCE', select_features=lt_inteira, search_distance='60000 Meters')
                arcpy.conversion.FeatureClassToFeatureClass(fc_select, os.path.join(gdb, 'Temas'),filename)
                arcpy.AddMessage(fr'Tema {fc} adicionado ao dataset "Temas"')
            elif 'UF' not in fields(fc):
                #make feature layer
                fc_layer = arcpy.management.MakeFeatureLayer(fc, 'fc_layer')
                fc_select = arcpy.management.SelectLayerByLocation(in_layer=fc_layer, overlap_type='WITHIN_A_DISTANCE', select_features=lt_inteira, search_distance='60000 Meters')
                arcpy.conversion.FeatureClassToFeatureClass(fc_select, os.path.join(gdb, 'Temas'),filename)
                arcpy.AddMessage(fr'Tema {fc} adicionado ao dataset "Temas"')
    arcpy.AddMessage('Temas adicionados ao geodatabase local')

def lista_dados_referenciais():
    filenames = [
        'Adutoras_SNIRH_ANA_2021',
        'Aerodromos_ANAC_2022',
        'Aerogeradores_ANEEL_2023',
        'APCB_Amazonia',
        'APCB_Caatinga',
        'APCB_Cerrado_Pantanal',
        'APCB_Mata_Atlantica',
        'APCB_ZonaCosteira',
        'Aves_Migratorias_Areas_Ameacadas_CEMAVE_2022',
        'Aves_Migratorias_Areas_Concentracao_CEMAVE_2022',
        'Bases_de_Combustíveis_EPE',
        'Bases_de_GLP_EPE',
        'Biomas_IBGE_2019_250000',
        'Blocos_Disponiveis_OPC_1009_ANP',
        'Cavidades_CANIE_19122022',
        'Centrais_Geradoras_Hidrelétricas_CGH_ANEEL',
        'CGH_Base_Existente_EPE',
        'CGH_Expansao_Planejada_EPE',
        'Conservacao_Aves_IBAS',
        'Dutos_de_escoamento_EPE',
        'Dutovias_MINFRA_2018',
        'EOL_Base_Existente_EPE',
        'EOL_Expansao_Planejada_EPE',
        'Ferrovias_MINFRA',
        'Gasodutos_de_distribuição_EPE_2023',
        'Gasodutos_de_transporte_EPE_2023',
        'Geologia_IBGE',
        'Geomorfologia_IBGE',
        'Hidrovias_ANTAQ',
        'IBAs_MataAtlantica_SaveBrasil_2023',
        'Lei_Mata_Atlantica_MMA',
        'Localidades_IBGE_2010',
        'LT_Existente_EPE_2023',
        'LT_Planejada_EPE_2023',
        'Municipios_IBGE_2022',
        'Ocorrencias_Fossiliferas_CPRM',
        'PCH_Base_Existente_EPE',
        'PCH_Expansao_Planejada_EPE_2023',
        'Pedologia_IBGE',
        'Pequenas_Centrais_Hidrelétricas_PCH_ANEEL',
        'Pivo_Central_Irrigacao_ANA_2019',
        'Plantas_de_biodiesel_EPE',
        'Plantas_de_etanol_EPE',
        'Polos_de_processamento_de_gás_natural_EPE',
        'Potencial_Cavidades_ICMBio',
        'Reserva_Biodiversidade_Mata_Atlantica_RBMA',
        'Rodovia_Estadual_DNIT',
        'Rodovia_Federal_DNIT',
        'RPPNs_ICMBio',
        'Sitios_Arqueologicos_IPHAN',
        'Subestações_Base_Existente_EPE_2023',
        'Subestacoes_Expansao_Planejada_EPE_2023',
        'Terminais_de_Petroleo_e_Derivados_EPE',
        'Territorios_Quilombolas_INCRA_2023',
        'Terras_Indigenas_FUNAI',
        'Trecho_Drenagem_ANA_2013',
        'UHE_Base_Existente_EPE_2023',
        'Unidades_de_Conservacao_MMA',
        'Usina_Fotovoltaica_UFV_ANEEL_2023',
        'Usina_Termeletricas_UTE_ANEEL_2023',
        'Usinas_Hidrelétricas_UHE_ANEEL_2023',
        'UTE_Biomassa_Existente_EPE_2023',
        'Vegetacao_IBGE',
        'Vilas_IBGE_2021',
        'Areas_Quilombolas_INCRA',
        'Areas_Urbanizadas_IBGE_2019',
        'Assentamentos_INCRA',
        'Processos_Minerarios_ANM']
    return filenames

def dissolve(fc):
    arcpy.env.overwriteOutput = True
    filename = os.path.basename(fc)
    fields_interesse = []
    if filename == 'Adutoras_SNIRH_ANA_2021':
        fields_interesse = ['adt_nm_adu','adt_status']
    elif filename == 'Aerodromos_ANAC_2022':
        fields_interesse = ['Codigo_OAC','CIAD','Denominaca']
    elif filename == 'Aerogeradores_ANEEL_2023':
        fields_interesse = ['NOME_EOL','DEN_AEG','POT_MW','CEG','OPERACAO']
    elif filename == 'APCB_Amazonia':
        fields_interesse = ['Import_bio','Prior_acao','COD_Area']
    elif filename == 'APCB_Caatinga':
        fields_interesse = ['Import_bio','Prior_acao','COD_area','Nome_area']
    elif filename == 'APCB_Cerrado_Pantanal':
        fields_interesse = ['Import_bio','Prior_acao','COD_area','NOME']
    elif filename == 'APCB_Mata_Atlantica':
        fields_interesse = ['ImportBio_','Prioridade','COD_area']
    elif filename == 'APCB_ZonaCosteira':
        fields_interesse = ['NOME_AP','IMP','PRIO']
    elif filename == 'Areas_Quilombolas_INCRA':
        fields_interesse = ['nm_comunid','nm_municip','responsave']
    elif filename == 'Areas_Urbanizadas_IBGE_2019':
        fields_interesse = ['Densidade','Tipo','Comparacao']
    elif filename == 'Assentamentos_INCRA':
        fields_interesse = ['nome_proje','municipio','capacidade','num_famili']
    elif filename == 'Aves_Migratorias_Areas_Ameacadas_CEMAVE_2022':
        fields_interesse = []#não tem campos de interesse, verificar se não vai ter bug
    elif filename == 'Aves_Migratorias_Areas_Concentracao_CEMAVE_2022':
        fields_interesse = []#não tem campos de interesse, verificar se não vai ter bug
    elif filename == 'Bases_de_Combustíveis_EPE':
        fields_interesse = ['nome_base','munic']
    elif filename == 'Bases_de_GLP_EPE':
        fields_interesse = ['nome_base','munic','razao_soci']
    elif filename == 'Biomas_IBGE_2019_250000':
        fields_interesse = ['Bioma']
    elif filename == 'Blocos_Disponiveis_OPC_1009_ANP':
        fields_interesse = ['nome_bacia','nomenclatu','nome_setor']
    elif filename == 'Cavidades_CANIE_2022':
        fields_interesse = ['Caverna','Municipio','Localidade']
    elif filename == 'Centrais_Geradoras_Hidrelétricas_CGH_ANEEL':
        fields_interesse = ['NOME']
    elif filename == 'CGH_Base_Existente_EPE':
        fields_interesse = ['NOME']
    elif filename == 'CGH_Expansao_Planejada_EPE':
        fields_interesse = ['Nome']
    elif filename == 'Conservacao_Aves_IBAS':
        fields_interesse = ['Nome_1']
    elif filename == 'Dutos_de_escoamento_EPE':
        fields_interesse = ['Nome_Dut_1','Categoria']
    elif filename == 'Dutovias_MINFRA_2018':
        fields_interesse = ['Nome_Duto']
    elif filename == 'EOL_Base_Existente_EPE':
        fields_interesse = ['Nome']
    elif filename == 'EOL_Expansao_Planejada_EPE':
        fields_interesse = ['nome']
    elif filename == 'Ferrovias_MINFRA':
        fields_interesse = ['tip_situac','bitola']
    elif filename == 'Gasodutos_de_distribuição_EPE_2023':
        fields_interesse = ['Distrib']
    elif filename == 'Gasodutos_de_transporte_EPE_2023':
        fields_interesse = ['Nome_Dut_1','Categoria']
    elif filename == 'Geologia_IBGE':
        fields_interesse = ['nm_unidade']
    elif filename == 'Geomorfologia_IBGE':
        fields_interesse = ['nm_unidade']
    elif filename == 'Hidrovias_ANTAQ':
        fields_interesse = ['HID_NM','HID_DS_CUR']
    elif filename == 'IBAs_MataAtlantica_SaveBrasil_2023':
        fields_interesse = ['Nome_1','Bioma']
    elif filename == 'Lei_Mata_Atlantica_MMA':
        fields_interesse = []#não tem campos de interesse, verificar se não vai ter bug
    elif filename == 'Localidades_IBGE_2010':
        fields_interesse = ['NM_LOCALID']
    elif filename == 'LT_Existente_EPE_2023':
        fields_interesse = ['Nome','Tensao','Ano_opera']
    elif filename == 'LT_Planejada_EPE_2023':
        fields_interesse = ['Nome','Tensao']
    elif filename == 'Municipios_IBGE_2022':
        fields_interesse = ['NM_MUN']
    elif filename == 'Ocorrencias_Fossiliferas_CPRM':
        fields_interesse = ['LOCALIDADE']
    elif filename == 'PCH_Base_Existente_EPE':
        fields_interesse = ['NOME']
    elif filename == 'PCH_Expansao_Planejada_EPE_2023':
        fields_interesse = ['nome']
    elif filename == 'Pedologia_IBGE':
        fields_interesse = ['legenda']
    elif filename == 'Pequenas_Centrais_Hidrelétricas_PCH_ANEEL':
        fields_interesse = ['NOME']
    elif filename == 'Pivo_Central_Irrigacao_ANA_2019':
        fields_interesse = ['NM_MUNICIP']
    elif filename == 'Plantas_de_biodiesel_EPE':
        fields_interesse = ['Nome']
    elif filename == 'Plantas_de_etanol_EPE':
        fields_interesse = ['Nome']
    elif filename == 'Polos_de_processamento_de_gás_natural_EPE':
        fields_interesse = ['Nome']
    elif filename == 'Potencial_Cavidades_ICMBio':
        fields_interesse = ['GRAU_DE_PO']
    elif filename == 'Reserva_Biodiversidade_Mata_Atlantica_RBMA':
        fields_interesse = ['CLASSE']
    elif filename == 'Rodovia_Estadual_DNIT':
        fields_interesse =['Unidade_Fe','Codigo_Rod']
    elif filename == 'Rodovia_Federal_DNIT':
        fields_interesse =['Codigo_BR']
    elif filename == 'RPPNs_ICMBio':
        fields_interesse = ['nome']
    elif filename == 'Sitios_Arqueologicos_IPHAN':
        fields_interesse = ['identifica']
    elif filename == 'Subestações_Base_Existente_EPE_2023':
        fields_interesse = ['Nome','Tensao','Ano_Opera']
    elif filename == 'Subestacoes_Expansao_Planejada_EPE_2023':
        fields_interesse = ['Nome','Tensao','Ano_Opera']
    elif filename == 'Terminais_de_Petroleo_e_Derivados_EPE':
        fields_interesse = ['nome_ter','munic']
    elif filename == 'Territorios_Quilombolas_INCRA_2023':
        fields_interesse = ['nm_comunid','nm_municip','nr_process','fase','responsave']
    elif filename == 'Terras_Indigenas_FUNAI':
        fields_interesse = ['terrai_nom','municipio_','etnia_nome']
    elif filename == 'Trecho_Drenagem_ANA_2013':
        fields_interesse = []#não tem campos de interesse, verificar se não vai ter bug
    elif filename == 'UHE_Base_Existente_EPE_2023':
        fields_interesse = ['NOME']
    elif filename == 'Unidades_de_Conservacao_MMA':
        fields_interesse =['NOME_UC1','CATEGORI3','ESFERA5','GRUPO4','ANO_CRIA6']
    elif filename == 'Usina_Fotovoltaica_UFV_ANEEL_2023':
        fields_interesse = ['nome','munic']
    elif filename == 'Usina_Termeletricas_UTE_ANEEL_2023':
        fields_interesse = ['nome']
    elif filename == 'Usinas_Hidrelétricas_UHE_ANEEL_2023':
        fields_interesse = ['NOME']
    elif filename == 'UTE_Biomassa_Existente_EPE_2023':
        fields_interesse = ['nome']
    elif filename == 'Vegetacao_IBGE':
        fields_interesse = ['legenda']
    elif filename == 'Vilas_IBGE_2021':
        fields_interesse = ['nome']
    elif filename == 'Processos_Minerarios_ANM':
        fields_interesse = ['PROCESSO','ANO','FASE','NOME','SUBS','AREA_HA','DSProcesso']

    if filename in temas_extra:
        field_split = fields_tema_extra.split(';')
        fields_interesse = field_split
        
    # Verifique se o campo "UF" existe no conjunto de features
    if 'UF' in [field.name for field in arcpy.ListFields(fc)]:
        fields_interesse.append('UF')


    # Caminho de saída para o conjunto de features dissolvido
    output_path = os.path.join(junkspace, f"{filename}_dissolved")
    output_path = arcpy.CreateUniqueName(output_path)
    
    if filename == 'Rodovia_Estadual_DNIT':
        arcpy.analysis.PairwiseBuffer(lt_inteira, os.path.join(junkspace, 'lt_buffer_DNIT1'), '60000 Meters')
        pw_clip = arcpy.analysis.PairwiseClip(fc, os.path.join(junkspace, 'lt_buffer_DNIT1'),  os.path.join(junkspace, 'pw_clip_DNIT1'))
        output = arcpy.analysis.PairwiseDissolve(pw_clip, output_path, fields_interesse)
    elif filename == 'Rodovia_Federal_DNIT':
        arcpy.analysis.PairwiseBuffer(lt_inteira, os.path.join(junkspace, 'lt_buffer_DNIT2'), '60000 Meters')
        pw_clip = arcpy.analysis.PairwiseClip(fc, os.path.join(junkspace, 'lt_buffer_DNIT2'), os.path.join(junkspace, 'pw_clip_DNIT2'))
        output = arcpy.analysis.PairwiseDissolve(pw_clip, output_path, fields_interesse)
    else:
        output = arcpy.analysis.PairwiseDissolve(fc, output_path, fields_interesse)

    return output, fields_interesse

def fields(fc):
    filename = os.path.basename(fc)
    fields_to_keep = []
    if filename == 'Adutoras_SNIRH_ANA_2021':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Aerodromos_ANAC_2022':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Aerogeradores_ANEEL_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'APCB_Amazonia':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'APCB_Caatinga':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'APCB_Cerrado_Pantanal':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'APCB_Mata_Atlantica':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'APCB_ZonaCosteira':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Areas_Quilombolas_INCRA':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Area','Extensao','OBS']
    elif filename == 'Areas_Urbanizadas_IBGE_2019':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Area','Extensao','OBS']
    elif filename == 'Assentamentos_INCRA':
        fields_to_keep = dissolve(fc)[1]+['Distancia','Vertices','Area','Extensao','OBS']
    elif filename == 'Aves_Migratorias_Areas_Ameacadas_CEMAVE_2022':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Aves_Migratorias_Areas_Concentracao_CEMAVE_2022':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Bases_de_Combustíveis_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Bases_de_GLP_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Biomas_IBGE_2019_250000':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Blocos_Disponiveis_OPC_1009_ANP':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Cavidades_CANIE_2022':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Centrais_Geradoras_Hidrelétricas_CGH_ANEEL':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'CGH_Base_Existente_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'CGH_Expansao_Planejada_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Conservacao_Aves_IBAS':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Dutos_de_escoamento_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'Dutovias_MINFRA_2018':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'EOL_Base_Existente_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'EOL_Expansao_Planejada_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Ferrovias_MINFRA':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'Gasodutos_de_distribuição_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'Gasodutos_de_transporte_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'Geologia_IBGE':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Geomorfologia_IBGE':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Hidrovias_ANTAQ':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'IBAs_MataAtlantica_SaveBrasil_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Lei_Mata_Atlantica_MMA':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Localidades_IBGE_2010':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'LT_Existente_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'LT_Planejada_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'Municipios_IBGE_2022':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area','Vertices']
    elif filename == 'Ocorrencias_Fossiliferas_CPRM':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'PCH_Base_Existente_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'PCH_Expansao_Planejada_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Pedologia_IBGE':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Pequenas_Centrais_Hidrelétricas_PCH_ANEEL':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Pivo_Central_Irrigacao_ANA_2019':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Plantas_de_biodiesel_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Plantas_de_etanol_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Polos_de_processamento_de_gás_natural_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Potencial_Cavidades_ICMBio':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Reserva_Biodiversidade_Mata_Atlantica_RBMA':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Rodovia_Estadual_DNIT':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'Rodovia_Federal_DNIT':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Vertices','Paralelism']
    elif filename == 'RPPNs_ICMBio':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Sitios_Arqueologicos_IPHAN':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Subestações_Base_Existente_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Subestacoes_Expansao_Planejada_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Terminais_de_Petroleo_e_Derivados_EPE':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Territorios_Quilombolas_INCRA_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Terras_Indigenas_FUNAI':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Trecho_Drenagem_ANA_2013':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Eixo_X','Eixo_Y']
    elif filename == 'UHE_Base_Existente_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Unidades_de_Conservacao_MMA':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Extensao','Area','Vertices','OBS']
    elif filename == 'Usina_Fotovoltaica_UFV_ANEEL_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Usina_Termeletricas_UTE_ANEEL_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Usinas_Hidrelétricas_UHE_ANEEL_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'UTE_Biomassa_Existente_EPE_2023':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Vegetacao_IBGE':
        fields_to_keep = dissolve(fc)[1]+['UF','Extensao','Area']
    elif filename == 'Vilas_IBGE_2021':
        fields_to_keep = dissolve(fc)[1]+['UF','Distancia','Vertices','Eixo_X','Eixo_Y']
    elif filename == 'Processos_Minerarios_ANM':
        fields_to_keep = dissolve(fc)[1]+['UF','Area','Extensao','Vertices']
    if filename in temas_extra:
        fd = fields_extras.split(';')
        fields_to_keep = dissolve(fc)[1]+list(fd)+['OBS']
        arcpy.AddMessage(fields_to_keep)
    if circuito_duplo == 'true':
        fields_to_keep.append('Circuito')
    return fields_to_keep

def ltxfeature(fc, lt):
    expression_geo = '!SHAPE.geodesicLength@KILOMETERS!'
    expression_proj = '!shape.length@kilometers!'
    dissolved_fc = dissolve(fc)[0]
    filename = os.path.basename(fc)
    if "Vertices" in fields(fc):
        if arcpy.Describe(fc).shapeType == 'Polyline' or arcpy.Describe(fc).shapeType == 'PolylineM' or arcpy.Describe(fc).shapeType == 'PolylineZ':    
            output_singlepart = arcpy.analysis.PairwiseIntersect(in_features=[lt, dissolved_fc], out_feature_class=os.path.join(junkspace,'LT_x_'+filename), output_type='POINT')
            output = arcpy.management.MultipartToSinglepart(output_singlepart, os.path.join(gdb_quantitativo,'LT_x_'+filename))
            arcpy.CalculateField_management(output, 'Eixo_X', '!shape.firstPoint.X!', 'PYTHON')
            arcpy.CalculateField_management(output, 'Eixo_Y', '!shape.firstPoint.Y!', 'PYTHON')
        elif arcpy.Describe(fc).shapeType == 'Polygon' or arcpy.Describe(fc).shapeType == 'PolygonM' or arcpy.Describe(fc).shapeType == 'PolygonZ':
            output = arcpy.analysis.PairwiseIntersect(in_features=[lt, dissolved_fc], out_feature_class=os.path.join(gdb_quantitativo,'LT_x_'+filename), output_type='LINE')
            arcpy.AddField_management(output, 'Area', 'FLOAT')
            if geodesic == 'true':
                arcpy.CalculateField_management(output, 'Extensao', expression_geo, 'PYTHON')
            else:
                arcpy.CalculateField_management(output, 'Extensao', expression_proj, 'PYTHON')
    elif "Vertices" not in fields(fc):
        if arcpy.Describe(fc).shapeType == 'Polyline' or arcpy.Describe(fc).shapeType == 'PolylineM' or arcpy.Describe(fc).shapeType == 'PolylineZ':
            lt=arcpy.analysis.PairwiseDissolve(in_features=lt, out_feature_class=os.path.join(junkspace,'dissolved_lt'))
            output_singlepart = arcpy.analysis.PairwiseIntersect(in_features=[lt, dissolved_fc], out_feature_class=os.path.join(junkspace,'LT_x_'+filename), output_type='POINT')
            output = arcpy.management.MultipartToSinglepart(output_singlepart, os.path.join(gdb_quantitativo,'LT_x_'+filename))
            arcpy.CalculateField_management(output, 'Eixo_X', '!shape.firstPoint.X!', 'PYTHON')
            arcpy.CalculateField_management(output, 'Eixo_Y', '!shape.firstPoint.Y!', 'PYTHON')
        elif arcpy.Describe(fc).shapeType == 'Polygon'or arcpy.Describe(fc).shapeType == 'PolygonM' or arcpy.Describe(fc).shapeType == 'PolygonZ':
            lt=arcpy.analysis.PairwiseDissolve(in_features=lt, out_feature_class=os.path.join(junkspace,'dissolved_lt'))
            output = arcpy.analysis.PairwiseIntersect(in_features=[lt, dissolved_fc], out_feature_class=os.path.join(gdb_quantitativo,'LT_x_'+filename), output_type='LINE')
            arcpy.AddField_management(output, 'Area', 'FLOAT')
            if geodesic == 'true':
                arcpy.CalculateField_management(output, 'Extensao', expression_geo, 'PYTHON')
            else:
                arcpy.CalculateField_management(output, 'Extensao', expression_proj, 'PYTHON')

def fxinteressexfeature(fc, fx_interesse):
    expression_proj_len = '!shape.length@kilometers!'
    expression_proj_area = '!shape.area@hectares!'
    expression_geo_area = '!SHAPE.geodesicArea@HECTARES!'
    expression_geo_len = '!SHAPE.geodesicLength@KILOMETERS!'
    dissolved_fc = dissolve(fc)[0]
    filename = os.path.basename(fc)
    if "Vertices" in fields(fc):
        if arcpy.Describe(fc).shapeType == 'Polyline' or arcpy.Describe(fc).shapeType == 'PolylineM' or arcpy.Describe(fc).shapeType == 'PolylineZ':
            output = arcpy.analysis.PairwiseIntersect([fx_interesse, dissolved_fc], os.path.join(gdb_quantitativo,'FxInteresse_x_'+filename))
            arcpy.AddField_management(output, 'Extensao', 'FLOAT')
            if geodesic == 'true':
                arcpy.AddMessage('Realizando cálculo de extensão geodésica')
                arcpy.CalculateField_management(output, 'Extensao', expression_geo_len, 'PYTHON')
            else:
                arcpy.CalculateField_management(output, 'Extensao', expression_proj_len, 'PYTHON')
        elif arcpy.Describe(fc).shapeType == 'Polygon' or arcpy.Describe(fc).shapeType == 'PolygonM' or arcpy.Describe(fc).shapeType == 'PolygonZ':
            output = arcpy.analysis.PairwiseIntersect([fx_interesse, dissolved_fc], os.path.join(gdb_quantitativo,'FxInteresse_x_'+filename))
            arcpy.AddField_management(output, 'Area', 'FLOAT')
            if geodesic == 'true':
                arcpy.AddMessage('Realizando cálculo de área geodésica')
                arcpy.CalculateField_management(output, 'Area', expression_geo_area, 'PYTHON')
            else:
                arcpy.CalculateField_management(output, 'Area', expression_proj_area, 'PYTHON')
    elif "Vertices" not in fields(fc):
        if arcpy.Describe(fc).shapeType == 'Polyline' or arcpy.Describe(fc).shapeType == 'PolylineM' or arcpy.Describe(fc).shapeType == 'PolylineZ':
            fx_interesse=arcpy.analysis.PairwiseDissolve(in_features=fx_interesse, out_feature_class=os.path.join(junkspace,'dissolved_fx'))
            output = arcpy.analysis.PairwiseIntersect([fx_interesse, dissolved_fc], os.path.join(gdb_quantitativo,'FxInteresse_x_'+filename))
            arcpy.AddField_management(output, 'Extensao', 'FLOAT')
            if geodesic == 'true':
                arcpy.AddMessage('Realizando cálculo de extensão geodésica')
                arcpy.CalculateField_management(output, 'Extensao', expression_geo_len, 'PYTHON')
            else:
                arcpy.CalculateField_management(output, 'Extensao', expression_proj_len, 'PYTHON')
        elif arcpy.Describe(fc).shapeType == 'Polygon' or arcpy.Describe(fc).shapeType == 'PolygonM' or arcpy.Describe(fc).shapeType == 'PolygonZ':
            fx_interesse=arcpy.analysis.PairwiseDissolve(in_features=fx_interesse, out_feature_class=os.path.join(junkspace,'dissolved_fx'))
            output = arcpy.analysis.PairwiseIntersect([fx_interesse, dissolved_fc], os.path.join(gdb_quantitativo,'FxInteresse_x_'+filename))
            arcpy.AddField_management(output, 'Area', 'FLOAT')
            if geodesic == 'true':
                arcpy.AddMessage('Realizando cálculo de área geodésica')
                arcpy.CalculateField_management(output, 'Area', expression_geo_area, 'PYTHON')
            else:
                arcpy.CalculateField_management(output, 'Area', expression_proj_area, 'PYTHON')


def ltnearfeature(fc,buffer,lt):
    dissolved_fc = dissolve(fc)[0]
    if geodesic == 'true':
        arcpy.analysis.Near(in_features = dissolved_fc, near_features = lt, search_radius = buffer, method = 'GEODESIC')
    else:
        arcpy.analysis.Near(in_features = dissolved_fc, near_features = lt, search_radius = buffer)
    expression = "round(!NEAR_DIST! / 1000.0, 3)"
    arcpy.CalculateField_management(dissolved_fc, 'Distancia', expression, "PYTHON")
    joinedfc = arcpy.management.JoinField(in_data=dissolved_fc, in_field='NEAR_FID', join_table=lt, join_field='OBJECTID')
    #arcpy.CalculateField_management(joinedfc, 'OBS', expression, "PYTHON")
    selectfc = arcpy.management.SelectLayerByAttribute(in_layer_or_view=joinedfc, selection_type="NEW_SELECTION", where_clause="NEAR_FID <> -1")
    output = arcpy.CopyFeatures_management(selectfc, fr'{gdb_quantitativo}\LT_Near_x_{os.path.basename(fc)}')
    arcpy.AddField_management(output, 'OBS', 'TEXT')
    expression = """def neardist(x):
        if x == 0:
            return 'Tema Sobrepõe a LT (Verificar planilha de sobreposição)'
        else:
            return ' '"""
    arcpy.management.CalculateField(in_table=output, field='OBS', expression='neardist(!NEAR_DIST!)', code_block=expression)

def tradutor(excel):
    dicionario = {
    'nm_adt_adu': 'Nome da adutora',
    'adt_status': 'Status',
    'adt_uf': 'UF da adutora',
    'Codigo_OAC': 'Código OAC',
    'CIAD': 'CIAD',
    'Denominaca': 'Denominação',
    'NOME_EOL': 'Nome',
    'DEN_AEG': 'Denominação',
    'POT_MW': 'Potência (mW)',
    'CEG': 'CEG',
    'OPERACAO': 'Operação',
    'Import_bio': 'Importância biológica',
    'Prior_acao': 'Prioridade da ação',
    'COD_Area': 'Código da área',
    'Nome_area': 'Nome da área',
    'NOME': 'Nome',
    'ImportBio': 'Importância biológica',
    'Prioridade': 'Prioridade',
    'COD_area': 'Código da área',
    'NOME_AP': 'Nome da APCB',
    'IMP': 'Importância',
    'PRIO': 'Prioridade',
    'nome_base': 'Nome da base',
    'munic': 'Município',
    'razao_soci': 'Razão social',
    'Bioma': 'Bioma',
    'nome_bacia': 'Nome da bacia',
    'nomenclatu': 'Nomenclatura',
    'nome_setor': 'Nome do setor',
    'Caverna': 'Caverna',
    'Municipio': 'Município',
    'Localidade': 'Localidade',
    'tip_situac': 'Tipo de situação',
    'bitola': 'Bitola',
    'Distrib': 'Distribuição',
    'Nome_Dut_1': 'Nome',
    'Categoria': 'Categoria',
    'nm_unidade': 'Nome da unidade',
    'HID_NM': 'Nome da hidrovia',
    'HID_DS_CUR': 'Descrição',
    'Nome_1': 'Nome',
    'NM_LOCALID': 'Nome da localidade',
    'Tensao': 'Tensão',
    'Ano_opera': 'Ano de operação',
    'NM_MUNICIP': 'Município',
    'SIGLA_UF': 'Sigla UF',
    'LOCALIDADE': 'Localidade',
    'Nome_Duto': 'Nome do duto',
    'NOME_ter': 'Nome',
    'GRAU_DE_PO': 'Grau de potencialidade',
    'CLASSE': 'Classe',
    'Tipo_Trech': 'Tipo de trecho',
    'Unidade_Fe': 'UF da rodovia',
    'Codigo_Rod': 'Código da rodovia',
    'Nome_Tipo': 'Nome',
    'Codigo_BR': 'Código da BR',
    'nome': 'Nome',
    'Identifica': 'Identificação',
    'nm_comunid': 'Nome da comunidade',
    'nm_municip': 'Município',
    'nr_process': 'Número do processo',
    'fase': 'Fase',
    'responsave': 'Responsável',
    'terrai_nom': 'Nome da Terra Indígena',
    'etnia_nome': 'Nome da etnia',
    'Extensao' : 'Extensão (km)',
    'Area' : 'Área (ha)',
    'Eixo_X' : 'Eixo X',
    'Eixo_Y' : 'Eixo Y',
    'Paralelism' : 'Paralelismo (Metros)',
    'Distancia' : 'Distância (km)',
    'Vertices' : 'Vértices',
    'OBS' : 'Observação',
    'identifica' : 'Identificação',
    'legenda' : 'Legenda',
    'NOME_UC1' : 'Nome da UC',
    'GRUPO4' : 'Grupo',
    'CATEGORI3' : 'Categoria',
    'ESFERA5' : 'Esfera',
    'ANO_CRIA6' : 'Ano de criação'}

    for row in excel.iter_rows():
        for cell in row:
            if cell.value in dicionario.keys():
                cell.value = dicionario[cell.value]
                

def toexcel(fc, related_field):
    campos_fc = [campo.name for campo in arcpy.ListFields(fc)]
    campos_related_field = fields(os.path.join(gdb_path, 'Temas', related_field))

    # Encontrar a interseção entre as colunas do fc e do related_field
    colunas_comuns = [campo for campo in campos_fc if campo in campos_related_field]

    if not colunas_comuns:
        arcpy.AddWarning(f"Não há colunas comuns entre {os.path.basename(fc)} e {related_field}")
        return

    # Remove duplicate fields
    colunas_comuns = list(set(colunas_comuns))

    # Obter os dados apenas das colunas comuns
    table = arcpy.da.TableToNumPyArray(fc, colunas_comuns, skip_nulls=False)
    df = pd.DataFrame(table)  # Inicializa a variável df com os dados da tabela

    if df.empty:
        df = pd.DataFrame({"Mensagem": ["Não há registros para esse tema na área de estudo"]})

    #lista todas as pastas na pasta quantitativo que não terminam com .xlsx
    lista_arquivos = [arquivo for arquivo in os.listdir(pasta_quantitativo) if not arquivo.endswith('.xlsx')]
    #lista todos os arquivos excel na pasta quantitativo


    if os.path.basename(fc).split('_x_')[1] in lista_arquivos:
        for arquivo in lista_arquivos:
            if os.path.basename(fc).split('_x_')[1] in arquivo:
                excel_saida = os.path.join(pasta_quantitativo, arquivo, f'{os.path.basename(fc)}.xlsx')
                #remove as colunas vazias
                df = df.dropna(axis=1, how='all')
                df.to_excel(excel_saida, index=False)
                arcpy.AddMessage(f'Planilha de quantitativo do tema {os.path.basename(fc)} gerado com sucesso!')
    else:
        os.mkdir(os.path.join(pasta_quantitativo, os.path.basename(fc).split('_x_')[1]))
        excel_saida = os.path.join(pasta_quantitativo, os.path.basename(fc).split('_x_')[1], f'{os.path.basename(fc)}.xlsx')
        #remove as colunas vazias
        df = df.dropna(axis=1, how='all')
        df.to_excel(excel_saida, index=False)
        arcpy.AddMessage(f'Planilha de quantitativo do tema {os.path.basename(fc)} gerado com sucesso!')

    excel = os.path.join(pasta_quantitativo,os.path.basename(fc).split('_x_')[1], os.path.basename(fc)+'.xlsx')
    workbook = op.load_workbook(excel)
    sheet = workbook.active
    #pinta a primeira linha de amarelo claro e negrito, mas não pinta além da última coluna
    for cell in sheet[1]:
        cell.fill = op.styles.PatternFill(start_color='eeb0ff', end_color='eeb0ff', fill_type='solid')
        cell.font = op.styles.Font(bold=True)
    #cria uma borda em volta da tabela
    border = op.styles.Border(left=op.styles.Side(border_style='thin', color='000000'),
                    right=op.styles.Side(border_style='thin', color='000000'),
                    top=op.styles.Side(border_style='thin', color='000000'),
                    bottom=op.styles.Side(border_style='thin', color='000000'))
    for row in sheet:
        for cell in row:
            cell.border = border
    #ajustar a largura da coluna pra 120 pixels cada coluna
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 30)
        sheet.column_dimensions[column].width = adjusted_width
    if os.path.basename(fc).split('_x_')[0] == 'LT_Near':
        sheet.insert_rows(1)
        if os.path.basename(fc).split("_x_")[1] == 'Unidades_de_Conservacao_MMA':
            sheet['A1'] = fr'Relação de distância do tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" com a LT - Raio de 50km'
        else:
            sheet['A1'] = fr'Relação de distância do tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" com a LT - Raio de 10km'
    elif os.path.basename(fc).split('_x_')[0] == 'LT':
        sheet.insert_rows(1)
        if arcpy.Describe(fc).shapeType == 'Polyline' or arcpy.Describe(fc).shapeType == 'PolylineM' or arcpy.Describe(fc).shapeType == 'PolylineZ':
            sheet['A1'] = fr'Extensão interceptada pelo tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" na Linha de Transmissão'
        elif arcpy.Describe(fc).shapeType == 'Multipoint' or arcpy.Describe(fc).shapeType == 'Point':
            sheet['A1'] = fr'Coordenadas dos pontos de interceptação do tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" na Linha de Transmissão'
    elif os.path.basename(fc).split('_x_')[0] == 'FxInteresse':
        sheet.insert_rows(1)
        if arcpy.Describe(fc).shapeType == 'Polyline' or arcpy.Describe(fc).shapeType == 'PolylineM' or arcpy.Describe(fc).shapeType == 'PolylineZ':
            sheet['A1'] = fr' Extensão interceptada pelo tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" na Faixa de Interesse'
        elif arcpy.Describe(fc).shapeType == 'Polygon' or arcpy.Describe(fc).shapeType == 'PolygonM' or arcpy.Describe(fc).shapeType == 'PolygonZ':
            sheet['A1'] = fr'Área interceptada pelo tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" na Faixa de Interesse'
    elif os.path.basename(fc).split('_x_')[0] == 'Paralelismo':
        sheet.insert_rows(1)
        sheet['A1'] = fr'Relação de Paralelismo do tema "{os.path.basename(fc).split("_x_")[1].replace("_"," ")}" com a LT - Foi considerado paralelismo as linhas dentro da Faixa de Interesse'	
    tradutor(sheet)
    #mescla as celulas da primeira linha até o final da tabela
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(sheet[1]))

    #quebra o texto da primeira linha e aumenta a altura da linha
    sheet.row_dimensions[1].height = 50
    #pinta de roxo escuro a primeira linha
    for cell in sheet[1]:
        cell.fill = op.styles.PatternFill(start_color='a570b3', end_color='a570b3', fill_type='solid')
        cell.font = op.styles.Font(bold=True)
        cell.alignment = op.styles.Alignment(horizontal='center', wrap_text=True, vertical = 'center')
        cell.border = border

    workbook.save(excel)


if atualizar_vao == 'true':
    #deleta todos os feature classes na pasta dados caruso
    arcpy.env.workspace = os.path.join(gdb_path, 'Dados_Caruso')
    arcpy.management.Delete(os.path.join(gdb_path, 'Dados_Caruso','Vao_LT'))
    arcpy.management.Delete(os.path.join(gdb_path, 'Dados_Caruso','Vao_FxInteresse'))
    if circuito_duplo == 'false':
        vao = criavao(lt_inteira, fx_interesse, vert_inicial)
        lt = vao[0]
        fx_interesse = vao[1]
    elif circuito_duplo == 'true':
        path = os.path.join(gdb_path, 'Dados_Caruso')
        #lista as linhas no shape de lt que tem circuito 1 e 2
        lt_circuito1 = arcpy.management.SelectLayerByAttribute(in_layer_or_view=lt_inteira, selection_type='NEW_SELECTION', where_clause="Circuito = 'C1'")
        lt_circuito2 = arcpy.management.SelectLayerByAttribute(in_layer_or_view=lt_inteira, selection_type='NEW_SELECTION', where_clause="Circuito = 'C2'")

        #LT
        lt_c1 = criavao(lt_circuito1, fx_interesse, vert_inicial)
        arcpy.management.Rename(lt_c1[0], 'Vao_LT_C1')
        lt_c2 = criavao(lt_circuito2, fx_interesse, vert_inicial)
        arcpy.management.Rename(lt_c2[0], 'Vao_LT_C2')
        lt = arcpy.management.Merge([os.path.join(path,'Vao_LT_C1'),os.path.join(path,'Vao_LT_C2')],os.path.join(path,'Vao_LT_Merged'))
        arcpy.management.Delete(os.path.join(path,'Vao_LT_C1'))
        arcpy.management.Delete(os.path.join(path,'Vao_LT_C2'))
        arcpy.management.Rename(os.path.join(path,'Vao_LT_Merged'), 'Vao_LT_Correto')

        #FXINTERESSE
        fx_c1 = criavao(lt_circuito1, fx_interesse, vert_inicial)
        arcpy.management.Rename(fx_c1[1], 'Vao_FxInteresse_C1')
        fx_c2 = criavao(lt_circuito2, fx_interesse, vert_inicial)
        arcpy.management.Rename(fx_c2[1], 'Vao_FxInteresse_C2')
        fx_interesse = arcpy.management.Merge([os.path.join(path,'Vao_FxInteresse_C1'),os.path.join(path,'Vao_FxInteresse_C2')],os.path.join(path,'Vao_FxInteresse_Merged'))
        arcpy.management.Delete(os.path.join(path,'Vao_FxInteresse_C1'))
        arcpy.management.Delete(os.path.join(path,'Vao_FxInteresse_C2'))
        arcpy.management.Rename(os.path.join(path,'Vao_FxInteresse_Merged'), 'Vao_FxInteresse_Correto')

        #CONSOLIDACAO
        arcpy.management.Delete(os.path.join(path,'Vao_LT'))
        arcpy.management.Rename(os.path.join(path,'Vao_LT_Correto'), 'Vao_LT')
        arcpy.management.Rename(os.path.join(path,'Vao_FxInteresse_Correto'), 'Vao_FxInteresse')
        lt = os.path.join(gdb_path, 'Dados_Caruso','Vao_LT')
        fx_interesse = os.path.join(gdb_path, 'Dados_Caruso','Vao_FxInteresse')
else: 
    lt = os.path.join(gdb_path, 'Dados_Caruso','Vao_LT')
    fx_interesse = os.path.join(gdb_path, 'Dados_Caruso','Vao_FxInteresse')

if atualizar == 'true':
    project(gdb_path, temas_extra)
else:
    pass

def parallel(fc,lt,fx_interesse):
    arcpy.MakeFeatureLayer_management(fc, 'fc_layer')
    arcpy.SelectLayerByLocation_management('fc_layer', 'INTERSECT', fx_interesse)
    arcpy.SplitLine_management('fc_layer', os.path.join(junkspace, 'split_lines'))
    arcpy.MakeFeatureLayer_management(os.path.join(junkspace, 'split_lines'), 'split_lines_layer')
    arcpy.SelectLayerByLocation_management('split_lines_layer', 'COMPLETELY_WITHIN', fx_interesse)
    arcpy.CalculateField_management('split_lines_layer', 'Paralelism', '1', 'PYTHON')
    arcpy.DeleteField_management('split_lines_layer', 'Extensao')
    identity_output = arcpy.analysis.Identity('split_lines_layer', fx_interesse, os.path.join(junkspace, 'split_lines_identity'))
    dissolve_output = arcpy.analysis.PairwiseDissolve(identity_output, os.path.join(junkspace, 'output_dissolve'),dissolve(os.path.basename(fc))[1]+['Paralelism','Vertices'])
    dissolve_output_layer = arcpy.MakeFeatureLayer_management(dissolve_output, 'dissolve_output_layer')
    lt_layer = arcpy.MakeFeatureLayer_management(lt, 'lt_layer')
    arcpy.SelectLayerByLocation_management(dissolve_output_layer, 'INTERSECT', lt_layer,'','NEW_SELECTION','INVERT')
    if geodesic == 'true':
        arcpy.CalculateField_management(dissolve_output_layer, 'Paralelism', '!SHAPE.geodesicLength@meters!', 'PYTHON')
    else:
        arcpy.CalculateField_management(dissolve_output_layer, 'Paralelism', '!shape.length@meters!', 'PYTHON')
    output_lines_final = os.path.join(junkspace,fr'Paralelismo_{os.path.basename(fc)}')
    arcpy.CopyFeatures_management(dissolve_output_layer, output_lines_final)
    arcpy.conversion.FeatureClassToFeatureClass(output_lines_final, os.path.join(gdb_path,'Quantitativo'), fr'Paralelismo_x_{os.path.basename(fc)}')
    

if temas_extra == '':
    arcpy.env.workspace = os.path.join(gdb_path, 'Temas')
    temas = arcpy.ListFeatureClasses()
    for tema in temas:
        if tema in lista_dados_referenciais():
            #conta quantos temas tem na pasta Temas
            count = len(temas)
            #faz um add mensage com o andamento do processo
            arcpy.AddMessage(f'Processando {tema} ({temas.index(tema)+1} de {count})')
            # Check if the feature class is "Unidade de Conservação"
            if os.path.basename(tema) == "Unidades_de_Conservacao_MMA":
                ltnearfeature(os.path.join(gdb_path, 'Temas', tema), '50000 Meters', lt)
            elif 'Distancia' in fields(tema) and os.path.basename(tema) != "Unidades_de_Conservacao_MMA":
                ltnearfeature(os.path.join(gdb_path, 'Temas', tema), '10000 Meters', lt)
            # Call ltxfeature and fxinteressexfeature functions
            ltxfeature(os.path.join(gdb_path, 'Temas', tema), lt)
            fxinteressexfeature(os.path.join(gdb_path, 'Temas', tema), fx_interesse)
            if 'Paralelism' in fields(tema):
                parallel(os.path.join(gdb_path, 'Temas', tema), lt, fx_interesse)
else:
    tema = os.path.basename(temas_extra)
    arcpy.env.workspace = os.path.join(gdb_path, 'Temas')
    if 'Distancia' in fields(temas_extra):
        ltnearfeature(os.path.join(gdb_path, 'Temas', tema), '10000 Meters', lt)
    if 'Paralelism' in fields(temas_extra):
        parallel(os.path.join(gdb_path, 'Temas', tema), lt, fx_interesse)
    
    # Call ltxfeature and fxinteressexfeature functions
    ltxfeature(os.path.join(gdb_path, 'Temas', tema), lt)
    fxinteressexfeature(os.path.join(gdb_path, 'Temas', tema), fx_interesse)


if temas_extra == '':
    arcpy.env.workspace = os.path.join(gdb_path, 'Quantitativo')
    temas = arcpy.ListFeatureClasses()
    for tema in temas:
        lastname = tema.split('_x_')[-1]
        arcpy.env.workspace = os.path.join(gdb_path, 'Temas')
        pasta_temas = arcpy.ListFeatureClasses()
        if lastname in lista_dados_referenciais() and lastname in pasta_temas:
            toexcel(os.path.join(gdb_path, 'Quantitativo', tema),lastname)
else:
    arcpy.env.workspace = os.path.join(gdb_path, 'Quantitativo')
    tema = os.path.basename(temas_extra)
    temas = arcpy.ListFeatureClasses()
    for fc in temas:
        if tema in fc:
            arcpy.AddMessage(f'Processando {fc} - TEMA EXTRA')
            name_fc = os.path.basename(fc)
            lastname = name_fc.split('_x_')[-1]
            toexcel(os.path.join(gdb_path, 'Quantitativo', fc),lastname)